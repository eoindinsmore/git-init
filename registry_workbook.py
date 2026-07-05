"""registry_workbook.py — the Excel control surface for the series registry.

    build   registry + store      -> registry_control.xlsx   (3 tabs)
    sync    registry_control.xlsx -> registry/*.yaml          (two-way write-back)

Charter note: ``registry/*.yaml`` stays the single source of truth. ``build`` mirrors it
into Excel (Tabs 1 & 3) and preserves your "New Series" input (Tab 2). ``sync`` writes
reviewed edits back — changed default transforms (Tab 1) and new series (Tab 2) — after
validating every row through ``SeriesSpec`` and preserving YAML comments (ruamel
round-trip). Run ``sync`` deliberately, then review the git diff and pull data for any new
series. The weekly scheduled task runs ``build`` only (a safe, YAML-read-only refresh).

Usage:
    python registry_workbook.py build     # refresh the workbook (default)
    python registry_workbook.py sync       # write workbook edits back into the registry
"""

from __future__ import annotations

import sys
from pathlib import Path
from typing import Any

import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

from quant import store
from quant import transforms as tf
from registry.loader import load_registry
from registry.schema import Category, Frequency, MacroTheme, SAStatus, SeriesSpec

REPO = Path(__file__).resolve().parent
WORKBOOK = REPO / "registry_control.xlsx"
REGISTRY_DIR = REPO / "registry"

SHEET_SERIES = "Series"
SHEET_NEW = "New Series"
SHEET_COVERAGE = "Coverage"

# Region buckets by country tag — mirrors app/data_access._REGION_BY_COUNTRY.
_REGION_BY_COUNTRY: dict[str, str] = {
    "US": "North America", "CA": "North America", "MX": "North America",
    "DE": "Europe", "GB": "Europe", "FR": "Europe", "EU": "Europe",
    "JP": "Asia", "CN": "Asia", "KR": "Asia", "IN": "Asia", "TW": "Asia",
    "CL": "Latin America", "PE": "Latin America",
}

SERIES_COLS = [
    "series_id", "source", "source_code", "name", "unit", "frequency", "sa_status",
    "metal", "country", "category", "macro_theme", "default_transformation", "caveats",
]
NEW_COLS = SERIES_COLS + ["source_params", "selector"]
COVERAGE_COLS = [
    "series_id", "name", "macro_theme", "region", "source", "frequency",
    "n_obs", "n_rows", "first_date", "last_date", "latest_as_of", "staleness_days",
]

_LABEL_TO_KIND = {label: kind for kind, label in tf.LABELS.items()}
_INK = "1A1A17"

# Dropdown option lists for data validation.
_DV_LISTS: dict[str, list[str]] = {
    "frequency": [f.value for f in Frequency],
    "sa_status": [s.value for s in SAStatus],
    "category": [c.value for c in Category],
    "macro_theme": [m.value for m in MacroTheme],
    "default_transformation": list(tf.LABELS.values()),
}


def _region_for(country: str | None) -> str:
    if not country:
        return "Global"
    return _REGION_BY_COUNTRY.get(country, country)


# --------------------------------------------------------------------------- build

def _series_rows() -> list[dict[str, Any]]:
    reg = load_registry()
    rows = []
    for sid in sorted(reg):
        spec = reg[sid]
        kind = tf.default_kind(spec.transformations)
        rows.append({
            "series_id": sid, "source": spec.source, "source_code": spec.source_code,
            "name": spec.name, "unit": spec.unit, "frequency": spec.frequency.value,
            "sa_status": spec.sa_status.value, "metal": spec.tags.metal or "",
            "country": spec.tags.country or "", "category": spec.tags.category.value,
            "macro_theme": spec.tags.macro_theme.value if spec.tags.macro_theme else "",
            "default_transformation": tf.LABELS[kind], "caveats": spec.caveats or "",
        })
    return rows


def _coverage_rows() -> list[dict[str, Any]]:
    facts = store.read_facts()
    reg = load_registry()
    if facts.empty:
        return []
    now = pd.Timestamp.now().normalize()
    rows = []
    for sid, g in facts.groupby("series_id"):
        spec = reg.get(sid)
        last_date = g["date"].max()
        theme = spec.tags.macro_theme.value if spec and spec.tags.macro_theme else ""
        rows.append({
            "series_id": sid, "name": spec.name if spec else sid, "macro_theme": theme,
            "region": _region_for(spec.tags.country if spec else None),
            "source": g["source"].iloc[0], "frequency": g["frequency"].iloc[0],
            "n_obs": int(g["date"].nunique()), "n_rows": int(len(g)),
            "first_date": g["date"].min().date(), "last_date": last_date.date(),
            "latest_as_of": g["as_of"].max().date(),
            "staleness_days": int((now - last_date).days),
        })
    return sorted(rows, key=lambda r: r["series_id"])


def _read_sheet(path: Path, sheet: str) -> list[dict[str, Any]]:
    """Read a sheet into a list of {header: value} dicts (blank cells -> '')."""
    if not path.exists():
        return []
    wb = load_workbook(path, data_only=True)
    if sheet not in wb.sheetnames:
        return []
    ws = wb[sheet]
    values = list(ws.values)
    if not values:
        return []
    headers = [str(h) if h is not None else "" for h in values[0]]
    out = []
    for row in values[1:]:
        record = {h: ("" if v is None else v) for h, v in zip(headers, row, strict=False)}
        if any(str(v).strip() for v in record.values()):
            out.append(record)
    return out


def _write_sheet(ws, columns: list[str], rows: list[dict[str, Any]]) -> None:
    ws.append(columns)
    for c in range(1, len(columns) + 1):
        cell = ws.cell(row=1, column=c)
        cell.font = Font(name="Arial", bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor=_INK)
        cell.alignment = Alignment(vertical="center")
    for row in rows:
        ws.append([row.get(col, "") for col in columns])
    ws.freeze_panes = "A2"
    for i, col in enumerate(columns, start=1):
        letter = get_column_letter(i)
        width = {"name": 46, "caveats": 60, "source_params": 24, "selector": 20}.get(col, 15)
        ws.column_dimensions[letter].width = width
        for r in range(2, ws.max_row + 1):
            ws.cell(row=r, column=i).font = Font(name="Arial")


def _add_validation(ws, columns: list[str], field: str, options: list[str], nrows: int) -> None:
    if field not in columns:
        return
    letter = get_column_letter(columns.index(field) + 1)
    dv = DataValidation(type="list", formula1='"' + ",".join(options) + '"', allow_blank=True)
    ws.add_data_validation(dv)
    dv.add(f"{letter}2:{letter}{max(nrows, 500)}")


def build(path: Path = WORKBOOK, new_series_rows: list[dict[str, Any]] | None = None) -> Path:
    """Write the 3-tab workbook from the registry + store.

    Tab 2 ("New Series") is populated from ``new_series_rows`` if given, otherwise the
    existing workbook's New Series rows are carried forward (so a refresh never eats
    pending input). Returns the path written.
    """
    if new_series_rows is None:
        new_series_rows = _read_sheet(path, SHEET_NEW)

    wb = Workbook()
    wb.remove(wb.active)

    ws_series = wb.create_sheet(SHEET_SERIES)
    series_rows = _series_rows()
    _write_sheet(ws_series, SERIES_COLS, series_rows)
    _add_validation(ws_series, SERIES_COLS, "default_transformation",
                    _DV_LISTS["default_transformation"], len(series_rows) + 1)

    ws_new = wb.create_sheet(SHEET_NEW)
    _write_sheet(ws_new, NEW_COLS, new_series_rows)
    for field, options in _DV_LISTS.items():
        _add_validation(ws_new, NEW_COLS, field, options, len(new_series_rows) + 1)

    ws_cov = wb.create_sheet(SHEET_COVERAGE)
    _write_sheet(ws_cov, COVERAGE_COLS, _coverage_rows())

    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)
    return path


# ---------------------------------------------------------------------------- sync

def _yaml():
    from ruamel.yaml import YAML

    y = YAML()
    y.preserve_quotes = True
    y.width = 100
    return y


def _yaml_path(source: str) -> Path:
    return REGISTRY_DIR / f"{source}.yaml"


def _flow_list(items: list[str]):
    """A ruamel sequence that dumps inline (``[yoy_pct]``), matching the repo style."""
    from ruamel.yaml.comments import CommentedSeq

    seq = CommentedSeq(items)
    seq.fa.set_flow_style()
    return seq


def _parse_kv(text: Any) -> dict[str, str]:
    """Parse a 'k=v;k=v' cell into a dict; '' -> {}."""
    s = str(text).strip()
    if not s:
        return {}
    out = {}
    for part in s.split(";"):
        if "=" in part:
            k, v = part.split("=", 1)
            out[k.strip()] = v.strip()
    return out


def _kind_from_label(label: Any) -> str:
    return _LABEL_TO_KIND.get(str(label).strip(), "level")


def _row_to_spec_dict(row: dict[str, Any]) -> dict[str, Any]:
    kind = _kind_from_label(row.get("default_transformation"))
    tags: dict[str, Any] = {"category": str(row.get("category", "")).strip()}
    if str(row.get("metal", "")).strip():
        tags["metal"] = str(row["metal"]).strip()
    if str(row.get("country", "")).strip():
        tags["country"] = str(row["country"]).strip()
    if str(row.get("macro_theme", "")).strip():
        tags["macro_theme"] = str(row["macro_theme"]).strip()
    spec: dict[str, Any] = {
        "series_id": str(row["series_id"]).strip(),
        "source": str(row["source"]).strip(),
        "source_code": str(row["source_code"]).strip(),
        "name": str(row["name"]).strip(),
        "unit": str(row["unit"]).strip(),
        "frequency": str(row["frequency"]).strip(),
        "sa_status": str(row["sa_status"]).strip(),
        "transformations": _flow_list([] if kind == "level" else [kind]),
        "tags": tags,
    }
    if str(row.get("caveats", "")).strip():
        spec["caveats"] = str(row["caveats"]).strip()
    if _parse_kv(row.get("source_params")):
        spec["source_params"] = _parse_kv(row.get("source_params"))
    if _parse_kv(row.get("selector")):
        spec["selector"] = _parse_kv(row.get("selector"))
    return spec


def _update_transform_in_yaml(source: str, series_id: str, kind: str) -> bool:
    """Set a series' ``transformations`` in its YAML file. Returns True if written."""
    path = _yaml_path(source)
    if not path.exists():
        return False
    y = _yaml()
    data = y.load(path.read_text(encoding="utf-8")) or []
    for item in data:
        if item.get("series_id") == series_id:
            item["transformations"] = _flow_list([] if kind == "level" else [kind])
            with path.open("w", encoding="utf-8") as fh:
                y.dump(data, fh)
            return True
    return False


def _append_series_to_yaml(source: str, spec_dict: dict[str, Any]) -> None:
    path = _yaml_path(source)
    y = _yaml()
    data = y.load(path.read_text(encoding="utf-8")) if path.exists() else None
    if data is None:
        data = []
    data.append(spec_dict)
    with path.open("w", encoding="utf-8") as fh:
        y.dump(data, fh)


def sync(path: Path = WORKBOOK) -> dict[str, list[str]]:
    """Write reviewed workbook edits back into the registry YAML.

    Applies changed default transforms (Series tab) and appends validated new series
    (New Series tab). Successful imports are rebuilt into the workbook (they move to the
    Series tab); rows that fail validation stay on the New Series tab with the reason.
    Returns {"changed": [...], "added": [...], "skipped": [...]}.
    """
    if not path.exists():
        raise FileNotFoundError(f"{path} not found — run `build` first.")

    reg = load_registry()
    report: dict[str, list[str]] = {"changed": [], "added": [], "skipped": []}

    # 1) default-transform edits from the Series tab
    for row in _read_sheet(path, SHEET_SERIES):
        sid = str(row.get("series_id", "")).strip()
        spec = reg.get(sid)
        if not spec:
            continue
        want = _kind_from_label(row.get("default_transformation"))
        changed = want != tf.default_kind(spec.transformations)
        if changed and _update_transform_in_yaml(spec.source, sid, want):
            report["changed"].append(f"{sid}: default -> {tf.LABELS[want]}")

    # 2) new series from the New Series tab
    failed_rows: list[dict[str, Any]] = []
    for row in _read_sheet(path, SHEET_NEW):
        sid = str(row.get("series_id", "")).strip()
        if not sid:
            continue
        if sid in reg:
            report["skipped"].append(f"{sid}: already in registry")
            failed_rows.append(row)
            continue
        try:
            spec_dict = _row_to_spec_dict(row)
            SeriesSpec.model_validate(spec_dict)  # loud validation before writing
        except Exception as e:  # noqa: BLE001 — surface the reason, keep the row for fixing
            report["skipped"].append(f"{sid}: {type(e).__name__} — {str(e).splitlines()[0]}")
            failed_rows.append(row)
            continue
        _append_series_to_yaml(spec_dict["source"], spec_dict)
        report["added"].append(f"{sid} -> registry/{spec_dict['source']}.yaml")

    # 3) regenerate the workbook; only failed new rows remain on the New Series tab
    build(path, new_series_rows=failed_rows)
    return report


# ---------------------------------------------------------------------------- main

def main(argv: list[str]) -> int:
    cmd = argv[1] if len(argv) > 1 else "build"
    if cmd == "build":
        out = build()
        print(f"built {out}  ({len(_series_rows())} series, {len(_coverage_rows())} with data)")
        return 0
    if cmd == "sync":
        report = sync()
        for key in ("changed", "added", "skipped"):
            for line in report[key]:
                print(f"  [{key}] {line}")
        n = len(report["changed"]) + len(report["added"])
        print(f"sync complete — {n} change(s) written; {len(report['skipped'])} skipped.")
        if report["added"]:
            print("Next: pull data for the new series (run its adapter), then review the git diff.")
        return 0
    print(f"unknown command {cmd!r}; use 'build' or 'sync'")
    return 2


if __name__ == "__main__":
    raise SystemExit(main(sys.argv))
