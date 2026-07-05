"""Write discovered candidates to a reviewable xlsx staging file.

The file has an ``include`` column (Y/N) so a human approves rows before they are
synced into the registry. Columns mirror the workbook's New Series tab plus review
aids (score, reason).
"""

from __future__ import annotations

from dataclasses import asdict
from pathlib import Path

import openpyxl
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

from discovery.relevance import Candidate

COLUMNS = [
    "include", "score", "reason", "series_id", "source", "source_code", "name",
    "unit", "frequency", "sa_status", "metal", "country", "category", "macro_theme",
    "caveats", "source_params", "selector",
]


def _fmt_kv(d: dict) -> str:
    return ";".join(f"{k}={v}" for k, v in d.items())


def stage_xlsx(candidates: list[Candidate], path: Path) -> Path:
    """Write candidates to ``path`` as one sheet per source, include=Y by default."""
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    by_source: dict[str, list[Candidate]] = {}
    for c in candidates:
        by_source.setdefault(c.source, []).append(c)

    header_font = Font(bold=True)
    header_fill = PatternFill("solid", fgColor="1F3A5F")
    for source, items in sorted(by_source.items()):
        ws = wb.create_sheet(source[:31])
        for j, col in enumerate(COLUMNS, 1):
            cell = ws.cell(1, j, col)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = header_fill
        for i, c in enumerate(items, 2):
            d = asdict(c)
            d["include"] = "Y" if c.include else "N"
            d["source_params"] = _fmt_kv(c.source_params)
            d["selector"] = _fmt_kv(c.selector)
            for j, col in enumerate(COLUMNS, 1):
                ws.cell(i, j, d.get(col, ""))
        # widths
        for j, col in enumerate(COLUMNS, 1):
            ws.column_dimensions[get_column_letter(j)].width = (
                40 if col in ("name", "reason", "caveats") else 14
            )
        ws.freeze_panes = "A2"
        _ = header_font  # (kept for parity with workbook styling)
    wb.save(path)
    return path
