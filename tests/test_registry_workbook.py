"""Tests for registry_workbook — the Excel control surface and its YAML write-back.

The write-back mutates the source-of-truth YAML, so it is tested carefully: comment
preservation, validation, transform edits, and new-series append (against a temp dir).
"""

from __future__ import annotations

import pytest
from ruamel.yaml import YAML

import registry_workbook as rw
from registry.schema import SeriesSpec

_FRED_YAML = """\
# FRED series declarations — keep this comment.
- series_id: us_industrial_production
  source: fred
  source_code: INDPRO
  name: "US Industrial Production"
  unit: "Index 2017=100"
  frequency: M
  sa_status: SA
  transformations: [yoy]
  tags:
    metal: null
    country: US
    category: activity
    macro_theme: activity
"""


@pytest.fixture
def temp_registry(tmp_path, monkeypatch):
    reg_dir = tmp_path / "registry"
    reg_dir.mkdir()
    (reg_dir / "fred.yaml").write_text(_FRED_YAML, encoding="utf-8")
    monkeypatch.setattr(rw, "REGISTRY_DIR", reg_dir)
    return reg_dir


def test_parse_kv():
    assert rw._parse_kv("geo=DE;nace_r2=B-D") == {"geo": "DE", "nace_r2": "B-D"}
    assert rw._parse_kv("") == {}
    assert rw._parse_kv("  ") == {}


def test_kind_from_label():
    assert rw._kind_from_label("YoY %") == "yoy_pct"
    assert rw._kind_from_label("Level") == "level"
    assert rw._kind_from_label("nonsense") == "level"


def test_row_to_spec_dict_validates():
    row = {
        "series_id": "us_cpi", "source": "fred", "source_code": "CPIAUCSL",
        "name": "US CPI", "unit": "Index", "frequency": "M", "sa_status": "SA",
        "metal": "", "country": "US", "category": "activity",
        "macro_theme": "inflation", "default_transformation": "YoY %",
        "caveats": "", "source_params": "", "selector": "",
    }
    spec_dict = rw._row_to_spec_dict(row)
    spec = SeriesSpec.model_validate(spec_dict)  # must not raise
    assert spec.tags.macro_theme.value == "inflation"
    assert spec.tags.metal is None  # blank -> omitted -> None
    assert spec_dict["transformations"] == ["yoy_pct"]


def test_update_transform_in_yaml_preserves_comment(temp_registry):
    assert rw._update_transform_in_yaml("fred", "us_industrial_production", "mom_pct")
    text = (temp_registry / "fred.yaml").read_text(encoding="utf-8")
    assert "keep this comment" in text  # ruamel round-trip kept the comment
    data = YAML().load(text)
    assert data[0]["transformations"] == ["mom_pct"]


def test_update_transform_to_level_empties_list(temp_registry):
    rw._update_transform_in_yaml("fred", "us_industrial_production", "level")
    data = YAML().load((temp_registry / "fred.yaml").read_text(encoding="utf-8"))
    assert list(data[0]["transformations"]) == []


def test_append_series_writes_valid_and_loadable(temp_registry):
    spec_dict = rw._row_to_spec_dict({
        "series_id": "us_cpi", "source": "fred", "source_code": "CPIAUCSL",
        "name": "US CPI", "unit": "Index", "frequency": "M", "sa_status": "SA",
        "metal": "", "country": "US", "category": "activity",
        "macro_theme": "inflation", "default_transformation": "YoY %",
        "caveats": "", "source_params": "", "selector": "",
    })
    rw._append_series_to_yaml("fred", spec_dict)
    text = (temp_registry / "fred.yaml").read_text(encoding="utf-8")
    assert "keep this comment" in text  # original preserved
    data = YAML().load(text)
    assert len(data) == 2
    # Every item still validates as a SeriesSpec (append produced a legal registry).
    for item in data:
        SeriesSpec.model_validate(item)


def test_build_writes_three_tabs(tmp_path):
    out = rw.build(path=tmp_path / "wb.xlsx")
    from openpyxl import load_workbook
    wb = load_workbook(out)
    assert wb.sheetnames == [rw.SHEET_SERIES, rw.SHEET_NEW, rw.SHEET_COVERAGE]
    # Series tab has one header row + one row per registry series.
    from registry.loader import load_registry
    assert wb[rw.SHEET_SERIES].max_row == len(load_registry()) + 1


def test_sync_roundtrip(tmp_path, monkeypatch):
    """End-to-end: edit a default (Series tab) + add valid & invalid new series."""
    from openpyxl import load_workbook

    from registry.loader import load_registry

    reg_dir = tmp_path / "registry"
    reg_dir.mkdir()
    (reg_dir / "fred.yaml").write_text(_FRED_YAML, encoding="utf-8")
    monkeypatch.setattr(rw, "REGISTRY_DIR", reg_dir)
    monkeypatch.setattr(rw, "load_registry", lambda: load_registry(reg_dir))

    wb = tmp_path / "wb.xlsx"
    valid = {
        "series_id": "us_cpi", "source": "fred", "source_code": "CPIAUCSL",
        "name": "US CPI", "unit": "Index", "frequency": "M", "sa_status": "SA",
        "metal": "", "country": "US", "category": "activity",
        "macro_theme": "inflation", "default_transformation": "YoY %",
        "caveats": "", "source_params": "", "selector": "",
    }
    invalid = {**valid, "series_id": "bad_one", "category": ""}  # category required -> rejected
    rw.build(path=wb, new_series_rows=[valid, invalid])

    # Edit the Series-tab default transform for the existing IP series -> "MoM %".
    w = load_workbook(wb)
    ws = w[rw.SHEET_SERIES]
    sid_col = rw.SERIES_COLS.index("series_id") + 1
    dt_col = rw.SERIES_COLS.index("default_transformation") + 1
    for r in range(2, ws.max_row + 1):
        if ws.cell(r, sid_col).value == "us_industrial_production":
            ws.cell(r, dt_col).value = "MoM %"
    w.save(wb)

    report = rw.sync(path=wb)

    data = YAML().load((reg_dir / "fred.yaml").read_text(encoding="utf-8"))
    by_id = {d["series_id"]: d for d in data}
    assert "us_cpi" in by_id  # valid new series appended
    assert list(by_id["us_industrial_production"]["transformations"]) == ["mom_pct"]
    assert "bad_one" not in by_id  # invalid never written
    assert any("us_cpi" in a for a in report["added"])
    assert any("bad_one" in s for s in report["skipped"])
    # Invalid row stays on the New Series tab; the imported one is gone.
    remaining = {r["series_id"] for r in rw._read_sheet(wb, rw.SHEET_NEW)}
    assert "bad_one" in remaining and "us_cpi" not in remaining


def test_build_preserves_new_series_rows(tmp_path):
    path = tmp_path / "wb.xlsx"
    draft = [{"series_id": "draft_x", "source": "fred", "source_code": "X", "name": "n",
              "unit": "u", "frequency": "M", "sa_status": "SA", "category": "activity",
              "macro_theme": "rates", "default_transformation": "Level"}]
    rw.build(path=path, new_series_rows=draft)
    rw.build(path=path)  # rebuild with no explicit rows -> should carry the draft forward
    rows = rw._read_sheet(path, rw.SHEET_NEW)
    assert any(r["series_id"] == "draft_x" for r in rows)
