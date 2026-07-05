"""LME COTR adapter — London Metal Exchange Commitments of Traders (MiFID II).

Charter note: rule #1 bans licensed LME *price* feeds. This is the **free MiFID II
positioning disclosure** (Article 58), ruled in-bounds as exchange-published
positioning data (rule #7). No price/settlement data is taken from LME here.

Access: the LME HTML pages are bot-protected (403), but the weekly XLSX files sit on
a CDN path that serves plain requests (HTTP 200). Filenames are fully predictable:

    https://www.lme.com/-/media/files/data/cotrs/<folder>/mifid-weekly-cotr-report--<key>--<DDMMYYYY>.xlsx

where <DDMMYYYY> is the Tuesday publication date. The adapter constructs these URLs
directly — no scraping of the listing page.

Each file is one metal's report. It carries the position date (prior Friday) AND the
actual publication timestamp, so point-in-time is exact:
    date    = position date (Friday close)
    as_of   = publication timestamp (Tuesday) — the honest "known by", no lookahead

One registry series selects one cell via source_params:
    folder   e.g. "ca-copper"          (CDN folder slug)
    key      e.g. "ca"                  (contract key in the filename)
    category e.g. "Investment Funds"    (MiFID II classification, substring match)
    side     "Long" | "Short"
    basis    "Total" (default) | "Risk Reducing" | "Other"
    weeks    how many recent Tuesdays to fetch per run (default 12)
"""

from __future__ import annotations

import io
from datetime import date, timedelta
from typing import Any

import pandas as pd
import requests

from adapters.base import AdapterError, BaseAdapter, TransientFetchError
from registry.schema import SeriesSpec

_CDN = "https://www.lme.com/-/media/files/data/cotrs"
_TIMEOUT = 60
_UA = {"User-Agent": "home-fund/0.1 (personal research)"}
_DEFAULT_WEEKS = 12


def _recent_tuesdays(today: date, weeks: int) -> list[date]:
    """Return the ``weeks`` most recent Tuesdays on/before ``today`` (newest first)."""
    # Monday=0 .. Sunday=6; Tuesday=1.
    offset = (today.weekday() - 1) % 7
    last_tue = today - timedelta(days=offset)
    return [last_tue - timedelta(weeks=i) for i in range(weeks)]


def _file_url(folder: str, key: str, tue: date) -> str:
    stamp = tue.strftime("%d%m%Y")
    return f"{_CDN}/{folder}/mifid-weekly-cotr-report--{key}--{stamp}.xlsx"


def read_cotr_xlsx(content: bytes) -> tuple[pd.Timestamp, pd.Timestamp, list[list]]:
    """Parse a COTR XLSX into (position_date, publish_ts, rows)."""
    import openpyxl

    wb = openpyxl.load_workbook(io.BytesIO(content), data_only=True)
    ws = wb[wb.sheetnames[0]]
    rows = [list(r) for r in ws.iter_rows(values_only=True)]
    # Row 2 = position date (Friday close); row 3 = publication timestamp (Tuesday).
    # The publication cell is tz-aware (UTC 'Z'); normalize both to tz-naive UTC.
    position_date = pd.to_datetime(rows[2][0], utc=True).tz_localize(None)
    publish_ts = pd.to_datetime(rows[3][0], utc=True).tz_localize(None)
    if pd.isna(position_date) or pd.isna(publish_ts):
        raise AdapterError("LME COTR: could not read position/publication dates (layout change)")
    return position_date, publish_ts, rows


def extract_position(rows: list[list], category: str, side: str, basis: str) -> float:
    """Extract one 'Number of Positions' cell for a category/side/basis."""
    # Locate the header row that names the MiFID categories, and its Long/Short row.
    hdr = next((i for i, r in enumerate(rows) if any(
        isinstance(c, str) and "Investment Funds" in c for c in r)), None)
    if hdr is None:
        raise AdapterError("LME COTR: category header row not found (layout change)")
    cat_col = next((j for j, c in enumerate(rows[hdr])
                    if isinstance(c, str) and category.lower() in c.lower()), None)
    if cat_col is None:
        raise AdapterError(f"LME COTR: category '{category}' not found in header")
    sub = rows[hdr + 1]
    want = side.lower()
    side_col = next((j for j in (cat_col, cat_col + 1)
                     if isinstance(sub[j], str) and sub[j].lower() == want), None)
    if side_col is None:
        raise AdapterError(f"LME COTR: side '{side}' not found under '{category}'")
    # Find the 'Number of Positions' block, then the requested basis row within it.
    start = next((i for i, r in enumerate(rows)
                  if isinstance(r[0], str) and "Number of Positions" in r[0]), None)
    if start is None:
        raise AdapterError("LME COTR: 'Number of Positions' block not found (layout change)")
    basis_row = None
    for i in range(start, min(start + 4, len(rows))):
        label = rows[i][2]
        if isinstance(label, str) and basis.lower() in label.lower():
            basis_row = i
            break
    if basis_row is None:
        raise AdapterError(f"LME COTR: basis '{basis}' not found in positions block")
    value = rows[basis_row][side_col]
    if value is None:
        raise AdapterError(f"LME COTR: empty cell for {category}/{side}/{basis}")
    return float(value)


class LmeCotrAdapter(BaseAdapter):
    source = "lme_cotr"

    def fetch_raw(self, spec: SeriesSpec) -> Any:
        p = spec.source_params
        folder = p.get("folder")
        key = p.get("key")
        if not folder or not key:
            raise AdapterError(
                f"LME COTR '{spec.series_id}': source_params needs 'folder' and 'key'"
            )
        weeks = int(p.get("weeks", _DEFAULT_WEEKS))
        today = pd.Timestamp.now(tz="UTC").date()
        files = []
        transient = 0
        # LME normally publishes on Tuesday, but the day shifts around UK bank
        # holidays (e.g. Wed after a bank-holiday Monday). For each week probe the
        # Tuesday and its neighbours, taking the first valid file for that week.
        for tue in _recent_tuesdays(today, weeks):
            for delta in (0, 1, -1, 2):  # Tue, Wed, Mon, Thu
                try:
                    r = requests.get(
                        _file_url(folder, key, tue + timedelta(days=delta)),
                        headers=_UA,
                        timeout=_TIMEOUT,
                    )
                except requests.RequestException:
                    transient += 1
                    continue
                # A real .xlsx is a ZIP archive (magic 'PK'). The CDN sometimes
                # returns HTTP 200 with an HTML soft-404 body — reject those.
                if r.status_code == 200 and r.content[:2] == b"PK":
                    files.append(r.content)
                    break  # got this week's report; stop probing neighbours
                if r.status_code == 429 or r.status_code >= 500:
                    transient += 1
        if not files:
            # Nothing at all fetched: if it was all network errors, let retry kick in.
            if transient:
                raise TransientFetchError(
                    f"LME COTR '{spec.series_id}': {transient} transient fetch failures"
                )
            raise AdapterError(
                f"LME COTR '{spec.series_id}': no report files found for the last {weeks} weeks "
                f"(check folder '{folder}'/key '{key}')"
            )
        return files

    def load_manual(self, spec: SeriesSpec) -> Any | None:
        """Manual-inbox fallback: read every .xlsx dropped in data/manual/lme_cotr/.

        Used to backfill deep history — LME's pre-2026 files use a different (`.xls`
        binary, legacy-named) scheme this adapter does not auto-fetch. Drop the
        historic .xlsx exports here and they are parsed like fetched files.
        """
        if not self._manual_dir.exists():
            return None
        files = [p.read_bytes() for p in sorted(self._manual_dir.glob("*.xlsx"))]
        return files or None

    def parse(self, spec: SeriesSpec, raw: Any) -> pd.DataFrame:
        p = spec.source_params
        category = p.get("category")
        side = p.get("side")
        basis = p.get("basis", "Total")
        if not category or not side:
            raise AdapterError(
                f"LME COTR '{spec.series_id}': source_params needs 'category' and 'side'"
            )
        rows_out = []
        for content in raw:
            position_date, publish_ts, sheet = read_cotr_xlsx(content)
            value = extract_position(sheet, category, side, basis)
            rows_out.append(
                {
                    "date": position_date,
                    "value": value,
                    "as_of": publish_ts,
                    "last_updated": publish_ts,
                }
            )
        return pd.DataFrame(rows_out, columns=["date", "value", "as_of", "last_updated"])
